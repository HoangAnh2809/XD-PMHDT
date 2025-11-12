"""
Report Generator - Excel and PDF Reports
"""
from datetime import datetime, date, timedelta
from typing import List, Dict
import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont


class ReportGenerator:
    
    @staticmethod
    def generate_excel_report(report_data: Dict, report_type: str) -> io.BytesIO:
        """Generate Excel report"""
        wb = Workbook()
        ws = wb.active
        
        # Set title
        ws.title = f"Báo cáo {report_type}"
        
        # Header styling
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title
        ws.merge_cells('A1:F1')
        title_cell = ws['A1']
        title_cell.value = f"BÁO CÁO {report_type.upper()}"
        title_cell.font = Font(bold=True, size=16, color="4472C4")
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Report info
        ws['A2'] = "Ngày tạo:"
        ws['B2'] = datetime.now().strftime("%d/%m/%Y %H:%M")
        ws['A3'] = "Từ ngày:"
        ws['B3'] = report_data.get('date_from', '')
        ws['A4'] = "Đến ngày:"
        ws['B4'] = report_data.get('date_to', '')
        
        # Summary section
        row = 6
        ws[f'A{row}'] = "TỔNG QUAN"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        
        row += 1
        summary = report_data.get('summary', {})
        for key, value in summary.items():
            ws[f'A{row}'] = key
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # Appointments table
        row += 2
        ws[f'A{row}'] = "CHI TIẾT LỊCH HẸN"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        
        row += 1
        headers = ['Mã', 'Khách hàng', 'Xe', 'Dịch vụ', 'Ngày', 'Trạng thái', 'Giá']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Data rows
        appointments = report_data.get('appointments', [])
        for appointment in appointments:
            row += 1
            data = [
                str(appointment.get('id', ''))[:8],
                appointment.get('customer_name', 'N/A'),
                appointment.get('vehicle_info', 'N/A'),
                appointment.get('service_type', 'N/A'),
                appointment.get('appointment_date', ''),
                appointment.get('status', ''),
                f"{appointment.get('actual_cost', 0):,.0f}"
            ]
            for col, value in enumerate(data, start=1):
                cell = ws.cell(row=row, column=col)
                cell.value = value
                cell.border = border
                if col == 7:  # Price column
                    cell.alignment = Alignment(horizontal='right')
        
        # Auto-adjust column widths
        for col in range(1, 8):
            max_length = 0
            column = get_column_letter(col)
            for cell in ws[column]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    @staticmethod
    def generate_pdf_report(report_data: Dict, report_type: str) -> io.BytesIO:
        """Generate PDF report"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#4472C4'),
            spaceAfter=30,
            alignment=1  # Center
        )
        
        title = Paragraph(f"BÁO CÁO {report_type.upper()}", title_style)
        elements.append(title)
        
        # Report info
        info_data = [
            ['Ngày tạo:', datetime.now().strftime("%d/%m/%Y %H:%M")],
            ['Từ ngày:', report_data.get('date_from', '')],
            ['Đến ngày:', report_data.get('date_to', '')],
        ]
        
        info_table = Table(info_data, colWidths=[2*inch, 4*inch])
        info_table.setStyle(TableStyle([
            ('FONT', (0, 0), (-1, -1), 'Helvetica', 10),
            ('FONT', (0, 0), (0, -1), 'Helvetica-Bold', 10),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ]))
        elements.append(info_table)
        elements.append(Spacer(1, 20))
        
        # Summary section
        summary_title = Paragraph("TỔNG QUAN", styles['Heading2'])
        elements.append(summary_title)
        elements.append(Spacer(1, 10))
        
        summary = report_data.get('summary', {})
        summary_data = [[key, str(value)] for key, value in summary.items()]
        
        summary_table = Table(summary_data, colWidths=[3*inch, 3*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#E7E6E6')),
            ('FONT', (0, 0), (0, -1), 'Helvetica-Bold', 10),
            ('FONT', (1, 0), (1, -1), 'Helvetica', 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 30))
        
        # Appointments table
        appointments_title = Paragraph("CHI TIẾT LỊCH HẸN", styles['Heading2'])
        elements.append(appointments_title)
        elements.append(Spacer(1, 10))
        
        # Table headers
        table_data = [['Mã', 'Khách hàng', 'Dịch vụ', 'Ngày', 'Trạng thái', 'Giá']]
        
        # Table rows
        appointments = report_data.get('appointments', [])
        for apt in appointments:
            table_data.append([
                str(apt.get('id', ''))[:8] + '...',
                apt.get('customer_name', 'N/A')[:20],
                apt.get('service_type', 'N/A')[:15],
                apt.get('appointment_date', '')[:10],
                apt.get('status', ''),
                f"{apt.get('actual_cost', 0):,.0f}"
            ])
        
        appointments_table = Table(table_data, colWidths=[0.8*inch, 1.5*inch, 1.3*inch, 1*inch, 1*inch, 1*inch])
        appointments_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONT', (0, 0), (-1, 0), 'Helvetica-Bold', 9),
            ('FONT', (0, 1), (-1, -1), 'Helvetica', 8),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('ALIGN', (1, 1), (1, -1), 'LEFT'),
            ('ALIGN', (-1, 1), (-1, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F2F2')])
        ]))
        elements.append(appointments_table)
        
        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        return buffer
